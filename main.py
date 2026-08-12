from fastapi import FastAPI, HTTPException, Query, Request, Form, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse, JSONResponse, Response, RedirectResponse, StreamingResponse
import pyotp
import qrcode
import io
import base64
import uvicorn
import hashlib
import secrets
import veritabani
import os
import html as html_lib
import sqlite3
import re
import zipfile
import json
import urllib.request
import urllib.parse
import hmac
from functools import lru_cache
from PIL import Image
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo

from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded

# İstek sınırlandırma motoru kurulumu
limiter = Limiter(key_func=get_remote_address, headers_enabled=True)
app = FastAPI(title="PDKS Ortak API Sistemi (Tam Güvenlikli)")
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

# CORS Güvenlik protokolleri tanımlaması
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

SIRKET_ANAHTARI = os.getenv("SIRKET_ANAHTARI", "BASE32SECRET3232QLDKSAJHGFRTYUIP")
ADMIN_SESSION_SECRET = os.getenv("ADMIN_SESSION_SECRET", SIRKET_ANAHTARI + "-admin-session")

def admin_oturum_tokeni_uret(kullanici_adi):
    son_kullanma = int(datetime.now().timestamp()) + 8 * 60 * 60
    veri = base64.urlsafe_b64encode(f"{kullanici_adi}|{son_kullanma}".encode()).decode().rstrip("=")
    imza = hmac.new(ADMIN_SESSION_SECRET.encode(), veri.encode(), hashlib.sha256).hexdigest()
    return f"{veri}.{imza}"

def admin_oturumu_gecerli(token):
    try:
        veri, imza = token.rsplit(".", 1)
        beklenen = hmac.new(ADMIN_SESSION_SECRET.encode(), veri.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(imza, beklenen): return False
        ham = base64.urlsafe_b64decode(veri + "=" * (-len(veri) % 4)).decode()
        _, son_kullanma = ham.rsplit("|", 1)
        return int(son_kullanma) >= int(datetime.now().timestamp())
    except Exception:
        return False

@app.middleware("http")
async def yonetici_api_korumasi(request: Request, call_next):
    korumali = request.url.path.startswith("/api/admin/") or request.url.path == "/api/get-logs"
    if korumali and not admin_oturumu_gecerli(request.cookies.get("pdks_admin_oturum", "")):
        return JSONResponse(status_code=401, content={"status":"error", "message":"Yönetici oturumu gerekli."})
    return await call_next(request)

PERSONEL_EK_ALANLARI = veritabani.PERSONEL_EK_ALANLARI

def tc_kimlik_gecerli(tc):
    tc = str(tc or "").strip()
    if re.fullmatch(r"TEST[0-9]{7}", tc.upper()):
        return veritabani.test_modu_acik_mi()
    if not re.fullmatch(r"[1-9][0-9]{10}", tc):
        return False
    rakamlar = [int(x) for x in tc]
    return ((sum(rakamlar[0:9:2]) * 7 - sum(rakamlar[1:8:2])) % 10 == rakamlar[9]
            and sum(rakamlar[:10]) % 10 == rakamlar[10])

def fotograf_bytes_hazirla(veri):
    if not veri:
        return None, None
    if len(veri) > 5 * 1024 * 1024:
        raise ValueError("Fotoğraf en fazla 5 MB olabilir.")
    try:
        resim = Image.open(io.BytesIO(veri)).convert("RGB")
        resim.thumbnail((600, 600))
        kenar = min(resim.size)
        sol = (resim.width - kenar) // 2
        ust = (resim.height - kenar) // 2
        resim = resim.crop((sol, ust, sol + kenar, ust + kenar)).resize((400, 400))
        cikti = io.BytesIO()
        resim.save(cikti, format="JPEG", quality=85, optimize=True)
        return "image/jpeg", base64.b64encode(cikti.getvalue()).decode("ascii")
    except Exception as exc:
        raise ValueError("Fotoğraf JPG, JPEG veya PNG biçiminde olmalıdır.") from exc

async def fotograf_hazirla(dosya):
    if not dosya or not getattr(dosya, "filename", ""):
        return None, None
    return fotograf_bytes_hazirla(await dosya.read())

def form_personel_alanlari(form):
    return {alan: str(form.get(alan, "")).strip() for alan in PERSONEL_EK_ALANLARI
            if alan not in ("foto_mime", "foto_base64")}

def sube_atamalarini_oku(form):
    try:
        veri = json.loads(str(form.get("sube_atamalari_json", "[]")) or "[]")
        if not isinstance(veri, list):
            raise ValueError
        return veri
    except (ValueError, TypeError, json.JSONDecodeError):
        raise ValueError("Ek şube bilgileri geçersiz gönderildi.")

def baslik_anahtari(deger):
    metin = str(deger or "").strip().lower().translate(str.maketrans("çğıöşü", "cgiosu"))
    return re.sub(r"[^a-z0-9]+", "_", metin).strip("_")

EXCEL_BASLIK_ESLEME = {
    "sicil": "sicil_no", "sicil_no": "sicil_no", "personel_no": "sicil_no",
    "tc": "tc_kimlik_no", "tc_kimlik_no": "tc_kimlik_no", "t_c_kimlik_no": "tc_kimlik_no",
    "ad": "isim", "isim": "isim", "soyad": "soyisim", "soyisim": "soyisim",
    "telefon": "telefon", "e_posta": "eposta", "eposta": "eposta", "email": "eposta",
    "departman": "departman", "gorev": "gorev", "unvan": "gorev", "maas": "maas",
    "sube": "sube_adi", "calisma_tipi": "calisma_modeli", "calisma_modeli": "calisma_modeli",
    "cinsiyet": "cinsiyet", "dogum_tarihi": "dogum_tarihi", "dogum_yeri": "dogum_yeri",
    "medeni_durum": "medeni_durum", "uyruk": "uyruk", "il": "il", "ilce": "ilce",
    "mahalle": "mahalle", "acik_adres": "acik_adres", "posta_kodu": "posta_kodu",
    "ise_giris_tarihi": "ise_giris_tarihi", "personel_turu": "personel_turu",
    "ogrenim_durumu": "ogrenim_durumu", "okul": "okul", "bolum": "bolum",
    "mezuniyet_yili": "mezuniyet_yili", "mezuniyet_durumu": "mezuniyet_durumu",
    "askerlik_durumu": "askerlik_durumu", "terhis_tarihi": "terhis_tarihi",
    "tecil_bitis_tarihi": "tecil_bitis_tarihi", "sgk_sicil_no": "sgk_sicil_no",
    "meslek_kodu": "meslek_kodu", "kan_grubu": "kan_grubu", "ehliyet_sinifi": "ehliyet_sinifi"
}

@lru_cache(maxsize=256)
def adres_api_getir(yol):
    url = "https://api.turkiyeapi.dev" + yol
    istek = urllib.request.Request(url, headers={"User-Agent": "PDKS/1.0"})
    with urllib.request.urlopen(istek, timeout=12) as cevap:
        return json.loads(cevap.read().decode("utf-8")).get("data", [])

def get_turkiye_timestamp():
    tz = ZoneInfo("Europe/Istanbul")
    return int(datetime.now(tz).timestamp())
@app.get("/api/get-qr")
@limiter.limit("30/minute")
def get_qr_code(request: Request):
    totp = pyotp.TOTP(SIRKET_ANAHTARI, interval=15)
    guncel_sifre = totp.now()

    qr = qrcode.QRCode(
        version=1, 
        error_correction=qrcode.constants.ERROR_CORRECT_L, 
        box_size=10, 
        border=4
    )
    qr.add_data(guncel_sifre)
    qr.make(fit=True)

    qr_resim = qr.make_image(fill_color="black", back_color="white")
    byte_arr = io.BytesIO()
    qr_resim.save(byte_arr, format='PNG')
    kodlanmis_resim = base64.b64encode(byte_arr.getvalue()).decode('utf-8')
    
    temiz_base64 = kodlanmis_resim.strip().replace(
        "\n", ""
    ).replace("\r", "")

    return JSONResponse(content={
        "status": "success",
        "qr_base64": f"data:image/png;base64,{temiz_base64}"
    })

@app.get("/api/admin/pdks-hatirlatmalar")
def admin_pdks_hatirlatmalar():
    try:
        veritabani.gelmeyen_personelleri_kontrol_et()
        gunluk = veritabani.gunluk_personel_durumlari()
        talepler = veritabani.duzeltme_talepleri_getir(tumu=False)
        kritik = [x for x in gunluk if x.get("durum") in (
            "İŞE GELMEDİ","EKSİK GİRİŞ","EKSİK ÇIKIŞ","PLAN HATASI"
        )]
        return JSONResponse(content={
            "status":"success",
            "sayi": len(kritik) + len(talepler),
            "gunluk": kritik,
            "talepler": talepler
        })
    except Exception as exc:
        return JSONResponse(status_code=500,content={"status":"error","message":str(exc)})

@app.get("/api/admin/gunluk-durum")
def admin_gunluk_durum(tarih: str = ""):
    try:
        hedef=tarih.strip() if tarih else None
        data=veritabani.gunluk_personel_durumlari(hedef)
        ozet={
            "toplam":len(data),
            "ise_gelmeyen":sum(1 for x in data if x["durum"]=="İŞE GELMEDİ"),
            "bugun_vardiyada":sum(1 for x in data if x["calisma_modeli"]=="VARDİYA" and x["planli_calisma"]),
            "esnek":sum(1 for x in data if x["calisma_modeli"]=="ESNEK"),
            "gec_gelen":0,
            "eksik_kayit":sum(1 for x in data if x["durum"] in ("EKSİK GİRİŞ","EKSİK ÇIKIŞ"))
        }
        # Geç gelen hesabı ilk giriş - plan başlangıcı ile yapılır.
        for x in data:
            if not x["ilk_giris"] or x["calisma_modeli"]=="ESNEK": continue
            p=veritabani.personel_sicil_ile_getir(x["sicil_no"])
            if not p: continue
            try:
                giris=datetime.strptime(x["tarih"]+" "+x["ilk_giris"],"%Y-%m-%d %H:%M:%S")
                bas=datetime.strptime(x["tarih"]+" "+str(p.get("mesai_baslangic") or "09:00")[:5],"%Y-%m-%d %H:%M")
                tol=int(p.get("personel_tolerans_dakika") or 20)
                if giris>bas+timedelta(minutes=tol):ozet["gec_gelen"]+=1
            except Exception: pass
        return JSONResponse(content={"status":"success","ozet":ozet,"data":data})
    except Exception as exc:
        return JSONResponse(status_code=500,content={"status":"error","message":str(exc)})

@app.get("/api/get-logs")
@limiter.limit("60/minute")
def get_logs(request: Request):
    try:
        ham_loglar = veritabani.tum_loglari_getir_api()
        formatli_loglar = []
        for log in ham_loglar:
            if isinstance(log, dict):
                durum = str(log.get("durum", "NORMAL"))
                formatli_loglar.append({
                    "log_id": log.get("id", "0"),
                    "personel": log.get(
                        "personel_ad_soyad", 
                        "Bilinmeyen Personel"
                    ),
                    "islem_turu": log.get("islem_turu", "GİRİŞ"),
                    "zaman": log.get("zaman_damgasi", "-"),
                    "sube": log.get("sube_adi", "Merkez"),
                    "durum_etiketi": durum if len(durum) > 6 else "NORMAL"
                })
        return JSONResponse(content={
            "status": "success", 
            "toplam_kayit": len(formatli_loglar), 
            "data": formatli_loglar
        })
    except Exception as e:
        return JSONResponse(content={
            "status": "error", 
            "message": f"Hata: {str(e)}"
        })
@app.post("/api/verify-camera-photo")
@limiter.limit("15/minute")
async def verify_camera_photo(
    request: Request,
    p_enlem: str = Form(...),
    p_boylam: str = Form(...),
    p_sapma: str = Form(...),
    cihaz_id: str = Form(...),
    cihaz_token: str = Form(...),
    zaman_damgasi: str = Form(...),
    okunan_qr_metni: str = Form(...),
    konum_sahte: str = Form("0"),
    konum_yasi_ms: str = Form("0"),
    konum_kaynagi: str = Form("web"),
    gelistirici_modu: str = Form("0"),
    vpn_proxy_aktif: str = Form("0"),
    android_guvenlik_surumu: str = Form("0"),
    kvkk_metin_surumu: str = Form(""),
    kvkk_bilgi_zamani: str = Form("")
):
    token_hash = hashlib.sha256(cihaz_token.encode()).hexdigest()
    personel = veritabani.personeli_cihazla_dogrula(cihaz_id, token_hash)
    if not personel:
        return JSONResponse(content={
            "status": "error", 
            "message": "Cihaz doğrulanamadı. Personel kurulumunu yeniden yapın."
        })
    if not personel.get("aktif"):
        veritabani.hata_logu_yaz(personel["id"], "QR", "PERSONEL_PASIF", "Personel hesabı pasif.")
        return JSONResponse(content={"status": "error", "message": "Personel hesabı pasif."})
    if not personel.get("sube_id"):
        veritabani.hata_logu_yaz(personel["id"], "QR", "SUBE_YOK", "Personele şube atanmamış.")
        return JSONResponse(content={"status": "error", "message": "Personele şube atanmamış."})

    if str(vpn_proxy_aktif).lower() in ("1", "true", "evet"):
        veritabani.hata_logu_yaz(personel["id"], "AĞ", "VPN_PROXY", "VPN veya proxy bağlantısı tespit edildi.")
        return JSONResponse(content={"status":"error", "message":"VPN/Proxy tespit edildi. Fake IP/VPN kapatıp tekrar deneyin."})

    # Güvenlik duvarı: kart basma yalnızca güncel Android native istemcisinden kabul edilir.
    # Tarayıcı/navigator.geolocation veya eski APK istemcisi Fake GPS kontrolünü atlayamaz.
    try:
        guvenlik_surumu = int(android_guvenlik_surumu or 0)
    except (TypeError, ValueError):
        guvenlik_surumu = 0
    if konum_kaynagi != "android-native" or guvenlik_surumu < 3:
        veritabani.hata_logu_yaz(personel["id"], "GPS", "GUVENLI_ISTEMCI_YOK", f"kaynak={konum_kaynagi}, surum={guvenlik_surumu}")
        return JSONResponse(content={
            "status":"error",
            "message":"Güvenli konum doğrulaması başarısız. PDKS uygulamasını güncelleyin; tarayıcıdan giriş/çıkış yapılamaz."
        })

    guncel_sunucu_zamani = get_turkiye_timestamp()
    try:
        gelen_zaman = int(zaman_damgasi)
    except ValueError:
        return JSONResponse(content={"status": "error", "message": "Zaman damgası tam sayı olmalıdır!"})

    if abs(guncel_sunucu_zamani - gelen_zaman) > 120:
        veritabani.hata_logu_yaz(personel["id"], "GPS", "ZAMAN_UYUSMAZLIGI", "Telefon ve sunucu zamanı uyuşmuyor.")
        return JSONResponse(content={
            "status": "error", 
            "message": "API Güvenlik Duvarı: Zaman aşımı!"
        })

    try:
        enlem_float = float(p_enlem) if p_enlem and p_enlem != "-" else 0.0
        boylam_float = float(p_boylam) if p_boylam and p_boylam != "-" else 0.0
        sapma_float = float(p_sapma) if p_sapma and p_sapma != "-" else 9999.0
        p_id_int = int(personel["id"])
        konum_yasi = int(float(konum_yasi_ms or 0))
    except ValueError:
        return JSONResponse(content={"status": "error", "message": "Veri formatı uyuşmazlığı!"})

    if not (-90 <= enlem_float <= 90 and -180 <= boylam_float <= 180) or (enlem_float == 0 and boylam_float == 0):
        veritabani.hata_logu_yaz(personel["id"], "GPS", "GPS_FORMAT", "Geçersiz koordinat gönderildi.")
        return JSONResponse(content={"status":"error", "message":"Geçersiz GPS koordinatı."})

    if str(konum_sahte).lower() in ("1", "true", "evet"):
        veritabani.hata_logu_yaz(personel["id"], "GPS", "FAKE_GPS", "Android sahte konum tespit etti.")
        return JSONResponse(content={"status":"error", "message":"Sahte konum tespit edildi. İşlem reddedildi."})

    # Geliştirici seçeneklerinin açık olması tek başına Fake GPS kanıtı değildir.
    # Bu bilgi yalnızca denetim/telemetri amacıyla tutulur; gerçek konum nesnesinin
    # Android mock bayrağı (konum_sahte) işlemi engelleyen sinyaldir.
    if konum_kaynagi == "android-native" and (konum_yasi < 0 or konum_yasi > 20000):
        veritabani.hata_logu_yaz(personel["id"], "GPS", "ESKI_KONUM", f"Konum yaşı: {konum_yasi}ms")
        return JSONResponse(content={"status":"error", "message":"Konum güncel değil. GPS'i açıp tekrar deneyin."})

    if sapma_float > 50.0 or sapma_float == 0.0:
        veritabani.hata_logu_yaz(personel["id"], "QR", "GPS_SAPMA", f"Konum sapması: {sapma_float}m")
        return JSONResponse(content={
            "status": "error", 
            "message": f"Konum güvenilir değil (Sapma: {sapma_float}m)!"
        })

    totp = pyotp.TOTP(SIRKET_ANAHTARI, interval=15)
    if not totp.verify(okunan_qr_metni, valid_window=1):
        veritabani.hata_logu_yaz(personel["id"], "QR", "QR_GECERSIZ", "Süresi dolmuş veya geçersiz karekod.")
        return JSONResponse(content={
            "status": "error", 
            "message": "Süresi dolmuş veya geçersiz karekod!"
        })

    try:
        basari_durumu, mesaj = veritabani.kart_basma_onayla(
            p_id=p_id_int, islem_turu=None, okunan_qr_sifresi=okunan_qr_metni,
            p_enlem=enlem_float, p_boylam=boylam_float, gelen_cihaz_id=cihaz_id
        )
        if not basari_durumu:
            veritabani.hata_logu_yaz(personel["id"], "QR", "KART_RED", mesaj)
        return JSONResponse(content={"status": "success" if basari_durumu else "error", "message": mesaj})
    except Exception as e:
        return JSONResponse(content={"status": "error", "message": f"Veritabanı Hatası: {str(e)}"})
@app.get("/pdks-ekran", response_class=HTMLResponse)
def pdks_ana_ekran():
    dosya_yolu = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), 
        "pdks_ekran.html"
    )
    with open(dosya_yolu, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())


KVKK_METIN_SURUMU = "KVKK-PDKS-2026-08-12-v1"

@app.get("/kvkk-aydinlatma", response_class=HTMLResponse)
def kvkk_aydinlatma(mobil: int = 0):
    firma = veritabani.firma_bilgilerini_getir()
    firma_adi = html_lib.escape(str(firma.get("firma_adi") or "İşveren / Veri Sorumlusu"))
    telefon = html_lib.escape(str(firma.get("telefon") or "İnsan Kaynakları birimi"))
    eposta = html_lib.escape(str(firma.get("eposta") or "İnsan Kaynakları birimi"))
    buton = """<button id='devam' disabled onclick='tamamla()'>Bilgi Edindim ve Devam Et</button><script>const cb=document.getElementById('okudum'),b=document.getElementById('devam');cb.addEventListener('change',()=>b.disabled=!cb.checked);function tamamla(){if(!cb.checked)return;if(window.Android&&typeof Android.kvkkBilgilendirmeyiTamamla==='function'){Android.kvkkBilgilendirmeyiTamamla()}else{location.href='/personel-kurulum'}}</script>""" if mobil else ""
    return HTMLResponse(f"""<!doctype html><html lang='tr'><head><meta charset='utf-8'><meta name='viewport' content='width=device-width,initial-scale=1'><title>PDKS KVKK Aydınlatma</title><style>body{{font-family:Arial,sans-serif;background:#f4f6f8;color:#17202a;margin:0}}main{{max-width:760px;margin:auto;padding:24px}}.kart{{background:white;border-radius:16px;padding:22px;box-shadow:0 4px 20px #0001}}h1{{font-size:24px}}h2{{font-size:18px;margin-top:24px}}p,li{{line-height:1.55}}.not{{background:#eef5ff;padding:12px;border-radius:10px}}label{{display:flex;gap:10px;margin:22px 0;font-weight:600}}button{{width:100%;padding:15px;border:0;border-radius:12px;font-weight:700}}button:disabled{{opacity:.45}}</style></head><body><main><div class='kart'><h1>Çalışan PDKS – KVKK Aydınlatma Metni</h1><p><b>Metin sürümü:</b> {KVKK_METIN_SURUMU}</p><p><b>Veri sorumlusu:</b> {firma_adi}<br><b>İletişim:</b> {eposta} / {telefon}</p><div class='not'>Bu ekran bir “açık rıza” talebi değildir. Amaç, kişisel veriler işlenmeden önce çalışanı bilgilendirmektir.</div><h2>İşlenen veriler ve amaç</h2><p>PDKS; personel/sicil bilgilerini, giriş-çıkış zamanı ve şube bilgisini, işlem anındaki konumu, cihaz eşleştirme kimliğini ve güvenlik kayıtlarını; mesai ve devam takibinin yürütülmesi, yetkili şubede bulunmanın doğrulanması, yetkisiz cihaz kullanımının ve sahte konum girişimlerinin önlenmesi amacıyla işler.</p><h2>Konum ve kamera</h2><p>Konum, PDKS giriş/çıkış doğrulaması sırasında kullanılır. Uygulama arka planda sürekli konum takibi yapmak üzere tasarlanmamıştır. Kamera yalnızca QR kod okutma işlevi için kullanılır; QR işlemi için fotoğraf arşivi oluşturulmaz.</p><h2>Toplama yöntemi ve hukuki sebep</h2><p>Veriler mobil uygulama, PDKS işlemleri ve güvenlik kayıtları üzerinden otomatik yollarla elde edilir. İşveren, her veri işleme faaliyeti için 6698 sayılı Kanun’daki somut işleme şartını kendi iş ilişkisi ve mevzuat yükümlülüklerine göre belirlemekle sorumludur; açık rıza gerekmeyen bir faaliyet rızaya bağlanmaz.</p><h2>Aktarım ve saklama</h2><p>Veriler yalnızca yetkili işveren birimleri ile, hukuken gerekli olduğu ölçüde yetkili kamu kurumları ve sistemin işletilmesinde görev alan veri işleyen hizmet sağlayıcılarla paylaşılabilir. Saklama süresi, ilgili mevzuatta öngörülen veya işleme amacı için gerekli süreyle sınırlı olmalıdır.</p><h2>Haklarınız</h2><p>6698 sayılı Kanun’un 11. maddesi kapsamındaki haklarınıza ilişkin taleplerinizi veri sorumlusuna yukarıdaki iletişim kanallarından iletebilirsiniz.</p><label><input id='okudum' type='checkbox'> Aydınlatma metnini okudum ve bilgi edindim.</label>{buton}</div></main></body></html>""")

@app.get("/personel-kurulum")
def personel_kurulum_ekrani():
    dosya_yolu = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), 
        "personel_kurulum.html"
    )
    if not os.path.exists(dosya_yolu):
        raise HTTPException(
            status_code=404, 
            detail="personel_kurulum.html bulunamadı!"
        )
    return FileResponse(dosya_yolu)

@app.get("/personel-sicil")
def personel_sicil_ekrani():
    dosya_yolu = os.path.join(os.path.dirname(os.path.abspath(__file__)), "personel_sicil.html")
    if not os.path.exists(dosya_yolu):
        raise HTTPException(status_code=404, detail="personel_sicil.html bulunamadı!")
    return FileResponse(dosya_yolu)

@app.get("/api/personel/oturum-kontrol")
@limiter.limit("30/minute")
def personel_oturum_kontrol(request: Request, cihaz_id: str = Query(...), cihaz_token: str = Query(...)):
    """Girişten hemen sonra cihaz+token eşleşmesini yan etkisiz doğrular.

    Ağır özet sorgularını çalıştırmaz; mobil giriş akışında yanlış/yarış sonucu
    oluşmuş token ile sicil kartına geçilmesini önler.
    """
    token_hash = hashlib.sha256(cihaz_token.encode()).hexdigest()
    personel = veritabani.personeli_cihazla_dogrula(cihaz_id, token_hash)
    if not personel or not personel.get("aktif"):
        return JSONResponse(status_code=401, content={"status": "error", "message": "Cihaz doğrulanamadı. Lütfen tekrar giriş yapın."})
    return JSONResponse(content={"status": "success"})

@app.get("/api/personel/ozet")
@limiter.limit("30/minute")
def personel_ozet(request: Request, cihaz_id: str = Query(...), cihaz_token: str = Query(...)):
    token_hash = hashlib.sha256(cihaz_token.encode()).hexdigest()
    personel = veritabani.personeli_cihazla_dogrula(cihaz_id, token_hash)
    if not personel or not personel.get("aktif"):
        return JSONResponse(status_code=401, content={"status": "error", "message": "Cihaz doğrulanamadı."})
    veritabani.gelmeyen_personelleri_kontrol_et()
    veri = veritabani.personel_mobil_ozeti(personel["id"], 30)
    veri["talepler"] = veritabani.duzeltme_talepleri_getir(personel_id=personel["id"], tumu=True)[:30]
    veri["amir_talepleri"] = veritabani.duzeltme_talepleri_getir(amir_personel_id=personel["id"], tumu=False)
    veri["durum_olaylari"] = veritabani.durum_olaylari_getir(personel_id=personel["id"])[:50]
    veri["amir_durum_olaylari"] = [
        x for x in veritabani.durum_olaylari_getir(sadece_bekleyen=True)
        if str(x.get("amir_personel_id") or "") == str(personel["id"])
    ]
    return JSONResponse(content={"status": "success", "data": veri})

def mobil_personeli_dogrula(cihaz_id, cihaz_token):
    return veritabani.personeli_cihazla_dogrula(cihaz_id, hashlib.sha256(cihaz_token.encode()).hexdigest())

@app.post("/api/personel/duzeltme-talebi")
async def personel_duzeltme_talebi(cihaz_id: str=Form(...), cihaz_token: str=Form(...), talep_turu: str=Form(...), istenen_zaman: str=Form(""), aciklama: str=Form("")):
    p=mobil_personeli_dogrula(cihaz_id,cihaz_token)
    if not p:return JSONResponse(status_code=401,content={"status":"error","message":"Cihaz doğrulanamadı."})
    ok,msg=veritabani.duzeltme_talebi_olustur(p["id"],talep_turu,istenen_zaman,aciklama)
    return JSONResponse(content={"status":"success" if ok else "error","message":msg})

@app.post("/api/personel/amir-karar")
async def personel_amir_karar(
    cihaz_id: str=Form(...), cihaz_token: str=Form(...), talep_id: str=Form(...),
    karar: str=Form(...), duzeltilmis_zaman: str=Form(""), aciklama: str=Form("")
):
    p=mobil_personeli_dogrula(cihaz_id,cihaz_token)
    if not p:
        return JSONResponse(status_code=401,content={"status":"error","message":"Cihaz doğrulanamadı."})
    return JSONResponse(
        status_code=403,
        content={
            "status":"error",
            "message":"PDKS kuralı: Amir giriş/çıkış saatini düzeltemez veya kesinleştiremez. Kayıt yönetici incelemesine gönderilmelidir."
        }
    )



@app.post("/api/personel/durum-olayi")
@limiter.limit("10/minute")
async def personel_durum_olayi(
    request: Request,
    cihaz_id: str=Form(...), cihaz_token: str=Form(...),
    olay_turu: str=Form(...), baslangic_tarihi: str=Form(...), bitis_tarihi: str=Form(...),
    aciklama: str=Form(""), belge: UploadFile|None=File(None)
):
    p=mobil_personeli_dogrula(cihaz_id,cihaz_token)
    if not p:
        return JSONResponse(status_code=401,content={"status":"error","message":"Cihaz doğrulanamadı."})
    tur=str(olay_turu or "").upper()
    belge_veri=None
    if belge and belge.filename:
        belge_veri=await belge.read()
        if len(belge_veri)>8*1024*1024:
            return JSONResponse(content={"status":"error","message":"Belge en fazla 8 MB olabilir."})
        izinli={"application/pdf","image/jpeg","image/png"}
        mime=(belge.content_type or "").lower()
        if mime not in izinli:
            return JSONResponse(content={"status":"error","message":"Yalnız PDF, JPG/JPEG veya PNG belge yüklenebilir."})
    if tur=="HASTALIK_RAPORU" and not belge_veri:
        return JSONResponse(content={"status":"error","message":"Hastalık raporu için PDF veya fotoğraf belge zorunludur."})

    ok,msg,olay_id=veritabani.durum_olayi_olustur(
        p["id"],tur,baslangic_tarihi,bitis_tarihi,aciklama,"PERSONEL"
    )
    if not ok:
        return JSONResponse(content={"status":"error","message":msg})
    if belge_veri and olay_id:
        b64=base64.b64encode(belge_veri).decode("ascii")
        sha=hashlib.sha256(belge_veri).hexdigest()
        if not veritabani.durum_olayi_belge_ekle(
            olay_id,p["id"],belge.filename,belge.content_type,b64,sha
        ):
            return JSONResponse(content={"status":"error","message":"Durum kaydedildi fakat belge saklanamadı. Yöneticiye bildirin."})
    return JSONResponse(content={"status":"success","message":msg,"olay_id":olay_id})

@app.post("/api/personel/amir-durum-karar")
async def personel_amir_durum_karar(
    cihaz_id: str=Form(...), cihaz_token: str=Form(...),
    olay_id: str=Form(...), karar: str=Form(...)
):
    p=mobil_personeli_dogrula(cihaz_id,cihaz_token)
    if not p:
        return JSONResponse(status_code=401,content={"status":"error","message":"Cihaz doğrulanamadı."})
    ok,msg=veritabani.amir_durum_karari(p["id"],olay_id,karar)
    return JSONResponse(content={"status":"success" if ok else "error","message":msg})

@app.get("/api/admin/durum-olaylari")
def admin_durum_olaylari():
    data=veritabani.durum_olaylari_getir()
    for x in data:
        x["belgeler"]=veritabani.olay_belgeleri_getir(x["olay_id"])
    return JSONResponse(content={"status":"success","data":data})

@app.post("/api/admin/durum-olayi-karar")
async def admin_durum_olayi_karar(
    olay_id: str=Form(...), karar: str=Form(...), aciklama: str=Form("")
):
    ok,msg=veritabani.yonetici_durum_karari(olay_id,karar,aciklama)
    return JSONResponse(content={"status":"success" if ok else "error","message":msg})

@app.post("/api/admin/manuel-durum")
async def admin_manuel_durum(
    personel_id: str=Form(...), tarih: str=Form(...), yeni_durum: str=Form(...), aciklama: str=Form(...)
):
    izinli={"İŞE GELMEDİ","RAPORLU","YILLIK İZİN","MAZERET İZNİ","ÜCRETSİZ İZİN","GÖREVLİ",
            "HAFTA TATİLİ","VARDİYA YOK","EKSİK GİRİŞ","EKSİK ÇIKIŞ","ÇALIŞTI"}
    if yeni_durum not in izinli:
        return JSONResponse(content={"status":"error","message":"Geçersiz manuel durum."})
    ok,msg=veritabani.yonetici_manuel_durum_kaydet(personel_id,tarih,yeni_durum,aciklama,"Yönetici")
    return JSONResponse(content={"status":"success" if ok else "error","message":msg})

@app.get("/api/admin/belge/{belge_id}")
def admin_belge_getir(belge_id: int):
    b=veritabani.personel_belgesi_getir(belge_id)
    if not b:
        raise HTTPException(status_code=404,detail="Belge bulunamadı.")
    return Response(
        content=base64.b64decode(b["dosya_base64"]),
        media_type=b["mime_turu"],
        headers={"Content-Disposition":f'inline; filename="{b["dosya_adi"]}"',"Cache-Control":"private, no-store"}
    )

def _puantaj_satirlari(baslangic,bitis,personel_id=None):
    b=datetime.strptime(baslangic,"%Y-%m-%d").date()
    s=datetime.strptime(bitis,"%Y-%m-%d").date()
    if s<b or (s-b).days>366:
        raise ValueError("Tarih aralığı 0-366 gün olmalıdır.")
    satirlar=[]
    gun=b
    while gun<=s:
        for x in veritabani.gunluk_personel_durumlari(gun.isoformat()):
            if personel_id is None or int(x["personel_id"])==int(personel_id):
                satirlar.append(x)
        gun+=timedelta(days=1)
    return satirlar

def _puantaj_excel_uret(satirlar,baslik):
    wb=Workbook();ws=wb.active;ws.title="Puantaj"
    kolonlar=["Tarih","Gün","Sicil","Ad Soyad","Şube","Çalışma Modeli","Vardiya",
              "Planlanan Giriş","Planlanan Çıkış","İlk Giriş","Son Çıkış","Net Çalışma",
              "Geç Kalma","Erken Çıkış","Durum","Kaynak","Açıklama"]
    ws.append(kolonlar)
    for c in ws[1]:
        c.font=Font(bold=True);c.alignment=Alignment(horizontal="center")
    gun_adlari=["Pazartesi","Salı","Çarşamba","Perşembe","Cuma","Cumartesi","Pazar"]
    for x in satirlar:
        dt=datetime.strptime(x["tarih"],"%Y-%m-%d").date()
        net=f'{x["toplam_dakika"]//60} sa {x["toplam_dakika"]%60} dk'
        ws.append([
            x["tarih"],gun_adlari[dt.weekday()],x["sicil_no"],x["personel"],x["sube"],
            x["calisma_modeli"],x["vardiya_grubu"],x["planlanan_giris"],x["planlanan_cikis"],
            x["ilk_giris"],x["son_cikis"],net,
            f'{x["gec_dakika"]} dk' if x["gec_dakika"] else "—",
            f'{x["erken_cikis_dakika"]} dk' if x["erken_cikis_dakika"] else "—",
            x["durum"],x["kaynak"],x["detay"]
        ])
    ws.freeze_panes="A2";ws.auto_filter.ref=ws.dimensions
    genislik=[13,13,12,28,22,16,14,16,16,13,13,18,13,13,20,16,45]
    for i,w in enumerate(genislik,1):
        ws.column_dimensions[chr(64+i)].width=w
    buf=io.BytesIO();wb.save(buf);buf.seek(0)
    return buf

@app.get("/api/admin/puantaj-excel")
def admin_puantaj_excel(
    baslangic: str=Query(...), bitis: str=Query(...), personel_id: int|None=Query(None)
):
    try:
        satirlar=_puantaj_satirlari(baslangic,bitis,personel_id)
        buf=_puantaj_excel_uret(satirlar,"Puantaj")
        ad=f"puantaj_{personel_id or 'tum_personel'}_{baslangic}_{bitis}.xlsx"
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition":f'attachment; filename="{ad}"'}
        )
    except Exception as exc:
        return JSONResponse(status_code=400,content={"status":"error","message":str(exc)})

@app.get("/yonetici-paneli", response_class=HTMLResponse)
def yonetici_paneli_arayuzu():
    if veritabani.ilk_kurulum_gerekli():
        return RedirectResponse("/ilk-kurulum", status_code=302)
    dosya_yolu = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), 
        "yonetici_paneli_gelismis.html"
    )
    try:
        with open(dosya_yolu, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        raise HTTPException(
            status_code=500, 
            detail=f"Hata: {dosya_yolu} sunucuda bulunamadı!"
        )

@app.get("/ilk-kurulum")
def ilk_kurulum_ekrani():
    if not veritabani.ilk_kurulum_gerekli():
        return RedirectResponse("/yonetici-paneli", status_code=302)
    return FileResponse(os.path.join(os.path.dirname(os.path.abspath(__file__)), "ilk_kurulum.html"))

@app.get("/api/kurulum-durumu")
def kurulum_durumu():
    return {"status": "success", "kurulum_gerekli": veritabani.ilk_kurulum_gerekli(), "veritabani": "PostgreSQL" if veritabani.POSTGRES_AKTIF else "SQLite (yerel test)"}

@app.get("/api/adres/iller")
def adres_illeri():
    try:
        data = adres_api_getir("/v2/provinces?fields=id,name&sort=name&limit=100")
        return {"status": "success", "data": data}
    except Exception:
        return JSONResponse(status_code=503, content={"status": "error", "message": "İl listesi alınamadı; adresi elle girebilirsiniz.", "data": []})

@app.get("/api/adres/ilceler/{il_id}")
def adres_ilceleri(il_id: int):
    try:
        data = adres_api_getir(f"/v2/districts?provinceId={il_id}&fields=id,name&sort=name&limit=1000")
        return {"status": "success", "data": data}
    except Exception:
        return JSONResponse(status_code=503, content={"status": "error", "message": "İlçe listesi alınamadı.", "data": []})

@app.get("/api/adres/mahalleler/{ilce_id}")
def adres_mahalleleri(ilce_id: int):
    try:
        data = adres_api_getir(f"/v2/neighborhoods?districtId={ilce_id}&fields=id,name,postalCode&sort=name&limit=1000")
        return {"status": "success", "data": data}
    except Exception:
        return JSONResponse(status_code=503, content={"status": "error", "message": "Mahalle listesi alınamadı; elle girebilirsiniz.", "data": []})

@app.post("/api/ilk-kurulum")
@limiter.limit("5/minute")
def ilk_kurulum(
    request: Request, firma_adi: str = Form(...), ad_soyad: str = Form(...),
    kullanici_adi: str = Form(...), sifre: str = Form(...), sifre_tekrar: str = Form(...),
    vergi_no: str = Form(""), telefon: str = Form(""), eposta: str = Form("")
):
    if sifre != sifre_tekrar:
        return JSONResponse(content={"status": "error", "message": "Parolalar eşleşmiyor."})
    basari, mesaj = veritabani.ilk_kurulumu_yap(firma_adi, ad_soyad, kullanici_adi, sifre, vergi_no, telefon, eposta)
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})

@app.post("/api/admin-login")
@limiter.limit("5/minute")
async def admin_login(
    request: Request,
    kullanici_adi: str = Form(...),
    sifre: str = Form(...)
):
    try:
        if veritabani.ilk_kurulum_gerekli():
            return JSONResponse(content={"status": "setup_required", "message": "Önce ilk firma kurulumunu tamamlayın."})
        if veritabani.yonetici_dogrula(kullanici_adi, sifre):
            cevap = JSONResponse(content={"status": "success", "message": "Giriş başarılı!"})
            cevap.set_cookie("pdks_admin_oturum", admin_oturum_tokeni_uret(kullanici_adi), max_age=28800, httponly=True, secure=True, samesite="strict")
            return cevap
        else:
            return JSONResponse(content={"status": "error", "message": "Kullanıcı adı veya şifre hatalı!"})
            
    except Exception as e:
        return JSONResponse(content={"status": "error", "message": f"Sistem Hatası: {str(e)}"})

@app.post("/api/admin-logout")
def admin_logout():
    cevap = JSONResponse(content={"status":"success"})
    cevap.delete_cookie("pdks_admin_oturum")
    return cevap
@app.get("/api/admin/personel-listesi")
async def api_personel_listesi():
    ham_personeller = veritabani.tum_personelleri_getir()
    formatli_personeller = []
    for p in ham_personeller:
        if isinstance(p, dict):
            kayit = {
                "id": str(p.get("id", "")),
                "isim": str(p.get("isim", "")),
                "soyisim": str(p.get("soyisim", "")),
                "departman": str(p.get("departman", "")),
                "maas": str(p.get("maas", "0")),
                "calisma_modeli": str(p.get("calisma_modeli", "SABİT")),
                "sicil_no": str(p.get("sicil_no") or ""),
                "telefon": str(p.get("telefon") or ""),
                "gorev": str(p.get("gorev") or ""),
                "sube_id": str(p.get("sube_id") or ""),
                "sube_adi": str(p.get("sube_adi") or "Şube Atanmamış"),
                "aktif": bool(p.get("aktif", 1)),
                "cihaz_id": str(p.get("cihaz_id") or "EŞLEŞMEDİ"),
                "foto_var": bool(p.get("foto_base64")),
                "foto_url": f"/api/personel/{p.get('id')}/foto" if p.get("foto_base64") else ""
            }
            for alan in PERSONEL_EK_ALANLARI:
                if alan not in ("foto_mime", "foto_base64"):
                    kayit[alan] = str(p.get(alan) or "")
            kayit["test_personeli"] = str(p.get("tc_kimlik_no") or "").upper().startswith("TEST")
            kayit["amir_id"] = str(veritabani.personel_amir_id_getir(p.get("id")) or "")
            kayit["sube_atamalari"] = veritabani.personel_subelerini_getir(p.get("id"))
            formatli_personeller.append(kayit)
    return JSONResponse(content={"status": "success", "data": formatli_personeller})

@app.get("/api/personel/{personel_id}/foto")
def personel_foto(personel_id: int):
    baglanti = veritabani.baglanti_ac()
    satir = baglanti.execute("SELECT foto_mime, foto_base64 FROM personeller WHERE id=?", (personel_id,)).fetchone()
    baglanti.close()
    if not satir or not satir[1]:
        raise HTTPException(status_code=404, detail="Fotoğraf bulunamadı.")
    return Response(
        content=base64.b64decode(satir[1]), media_type=satir[0] or "image/jpeg",
        headers={"Cache-Control": "private, max-age=300"}
    )

@app.post("/api/admin/personel-ekle")
async def api_personel_ekle(request: Request):
    form = await request.form()
    try:
        sube_atamalari = sube_atamalarini_oku(form)
    except ValueError as exc:
        return JSONResponse(content={"status": "error", "message": str(exc)})
    tc = str(form.get("tc_kimlik_no", "")).strip()
    if not tc_kimlik_gecerli(tc):
        return JSONResponse(content={"status": "error", "message": "Geçerli 11 haneli TC kimlik numarası zorunludur."})
    try:
        foto_mime, foto_base64 = await fotograf_hazirla(form.get("foto"))
    except ValueError as exc:
        return JSONResponse(content={"status": "error", "message": str(exc)})
    if not foto_base64:
        return JSONResponse(content={"status": "error", "message": "Personel fotoğrafı zorunludur."})
    ek = form_personel_alanlari(form)
    ek["tc_kimlik_no"] = tc.upper() if tc.upper().startswith("TEST") else tc
    if not all(str(form.get(a, "")).strip() for a in ("isim", "soyisim", "sicil_no", "il", "ilce", "acik_adres")):
        return JSONResponse(content={"status": "error", "message": "Ad, soyad, sicil numarası, il, ilçe ve açık adres zorunludur."})
    if ek.get("cinsiyet") == "Erkek" and not ek.get("askerlik_durumu"):
        return JSONResponse(content={"status": "error", "message": "Erkek personel için askerlik durumu zorunludur."})
    ek.update({"foto_mime": foto_mime, "foto_base64": foto_base64})
    basari, mesaj = veritabani.personel_ekle(
        form.get("isim", ""), form.get("soyisim", ""), form.get("departman", ""),
        form.get("maas", "0"), form.get("calisma_modeli", "SABİT"),
        form.get("sicil_no", ""), form.get("telefon", ""), form.get("gorev", ""),
        form.get("sube_id", ""), form.get("aktif", "1"), **ek
    )
    if basari:
        personel = veritabani.personel_sicil_ile_getir(str(form.get("sicil_no", "")).strip())
        if personel:
            basari, sube_mesaji = veritabani.personel_subelerini_ayarla(
                personel["id"], form.get("sube_id", ""), sube_atamalari
            )
            if not basari:
                mesaj = "Personel eklendi ancak şube yetkileri kaydedilemedi: " + sube_mesaji
            else:
                basari, amir_mesaji = veritabani.personel_amir_ata(personel["id"], form.get("amir_id", ""))
                if not basari:
                    mesaj = "Personel eklendi ancak ilgili amir kaydedilemedi: " + amir_mesaji
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})

@app.post("/api/admin/personel-excel-aktar")
async def personel_excel_aktar(
    excel: UploadFile = File(...), fotograflar: UploadFile = File(...), onay: str = Form("0")
):
    if not (excel.filename or "").lower().endswith(".xlsx"):
        return JSONResponse(content={"status": "error", "message": "Excel dosyası .xlsx biçiminde olmalıdır."})
    try:
        excel_veri = await excel.read()
        zip_veri = await fotograflar.read()
        if len(excel_veri) > 15 * 1024 * 1024 or len(zip_veri) > 100 * 1024 * 1024:
            raise ValueError("Yüklenen dosya boyutu sınırı aşıldı.")
        wb = load_workbook(io.BytesIO(excel_veri), read_only=True, data_only=True)
        ws = wb.active
        satirlar = ws.iter_rows(values_only=True)
        basliklar = next(satirlar, None)
        if not basliklar:
            raise ValueError("Excel dosyası boş.")
        alanlar = [EXCEL_BASLIK_ESLEME.get(baslik_anahtari(x), "") for x in basliklar]
        zorunlu = {"sicil_no", "tc_kimlik_no", "isim", "soyisim"}
        if not zorunlu.issubset(set(alanlar)):
            raise ValueError("Excel'de Sicil No, TC Kimlik No, Ad ve Soyad sütunları zorunludur.")
        foto_haritasi = {}
        with zipfile.ZipFile(io.BytesIO(zip_veri)) as zf:
            toplam = sum(x.file_size for x in zf.infolist())
            if toplam > 150 * 1024 * 1024:
                raise ValueError("Fotoğraf ZIP içeriği çok büyük.")
            for bilgi in zf.infolist():
                ad = os.path.basename(bilgi.filename)
                kok, uzanti = os.path.splitext(ad)
                if kok and uzanti.lower() in (".jpg", ".jpeg", ".png"):
                    foto_haritasi[kok.strip().upper()] = zf.read(bilgi)

        subeler = {str(x.get("sube_adi", "")).strip().upper(): x.get("id") for x in veritabani.tum_subeleri_getir()}
        dbc = veritabani.baglanti_ac()
        mevcut_siciller = {str(x[0]).strip().upper() for x in dbc.execute("SELECT sicil_no FROM personeller WHERE sicil_no IS NOT NULL")}
        mevcut_tc = {str(x[0]).strip() for x in dbc.execute("SELECT tc_kimlik_no FROM personeller WHERE tc_kimlik_no IS NOT NULL")}
        dbc.close()
        sonuclar, gecerli_kayitlar = [], []
        gorulen_tc, gorulen_sicil = set(), set()
        for no, hucreler in enumerate(satirlar, start=2):
            kayit = {alanlar[i]: ("" if v is None else str(v).strip()) for i, v in enumerate(hucreler) if i < len(alanlar) and alanlar[i]}
            if not any(kayit.values()):
                continue
            hatalar = []
            sicil = kayit.get("sicil_no", "").upper()
            tc = kayit.get("tc_kimlik_no", "")
            if not sicil: hatalar.append("Sicil numarası boş")
            if not tc_kimlik_gecerli(tc): hatalar.append("TC kimlik numarası geçersiz")
            if not kayit.get("isim"): hatalar.append("Ad boş")
            if not kayit.get("soyisim"): hatalar.append("Soyad boş")
            if not kayit.get("il"): hatalar.append("İl boş")
            if not kayit.get("ilce"): hatalar.append("İlçe boş")
            if not kayit.get("acik_adres"): hatalar.append("Açık adres boş")
            if kayit.get("cinsiyet") == "Erkek" and not kayit.get("askerlik_durumu"):
                hatalar.append("Erkek personel için askerlik durumu boş")
            if sicil in gorulen_sicil: hatalar.append("Excel içinde sicil tekrarı")
            if tc in gorulen_tc: hatalar.append("Excel içinde TC tekrarı")
            if sicil in mevcut_siciller: hatalar.append("Sicil numarası sistemde kayıtlı")
            if tc in mevcut_tc: hatalar.append("TC kimlik numarası sistemde kayıtlı")
            gorulen_sicil.add(sicil); gorulen_tc.add(tc)
            foto = foto_haritasi.get(sicil)
            if not foto: hatalar.append(f"{sicil}.jpg/png fotoğrafı bulunamadı")
            sube_adi = kayit.pop("sube_adi", "").upper()
            if sube_adi and sube_adi not in subeler: hatalar.append("Şube sistemde bulunamadı")
            if not hatalar:
                foto_mime, foto_base64 = fotograf_bytes_hazirla(foto)
                kayit.update({"sube_id": subeler.get(sube_adi), "foto_mime": foto_mime, "foto_base64": foto_base64})
                gecerli_kayitlar.append(kayit)
            sonuclar.append({"satir": no, "sicil_no": sicil, "ad_soyad": f"{kayit.get('isim','')} {kayit.get('soyisim','')}".strip(), "hatalar": hatalar})

        if onay == "1" and any(x["hatalar"] for x in sonuclar):
            return JSONResponse(content={"status": "error", "message": "Hatalı satırlar düzeltilmeden aktarım yapılamaz.", "data": sonuclar})
        eklenen = 0
        if onay == "1":
            for k in gecerli_kayitlar:
                temel = {a: k.pop(a, "") for a in ("isim", "soyisim", "departman", "maas", "calisma_modeli", "sicil_no", "telefon", "gorev", "sube_id")}
                basari, mesaj = veritabani.personel_ekle(
                    temel["isim"], temel["soyisim"], temel["departman"], temel["maas"] or 0,
                    temel["calisma_modeli"] or "SABİT", temel["sicil_no"], temel["telefon"],
                    temel["gorev"], temel["sube_id"], 1, **k
                )
                if not basari:
                    raise ValueError(f"{temel['sicil_no']}: {mesaj}")
                eklenen += 1
        return JSONResponse(content={"status": "success", "message": f"{eklenen} personel aktarıldı." if onay == "1" else "Önizleme hazır.", "data": sonuclar, "gecerli": len(gecerli_kayitlar), "hatali": sum(bool(x['hatalar']) for x in sonuclar)})
    except Exception as exc:
        return JSONResponse(content={"status": "error", "message": str(exc)})

@app.post("/api/admin/personel-sil")
async def api_personel_sil(personel_id: str = Form(...)):
    basari, mesaj = veritabani.personel_sil(personel_id)
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})

@app.get("/api/admin/sube-listesi")
async def api_sube_listesi():
    try:
        ham_subeler = veritabani.tum_subeleri_getir()
        formatli_subeler = []
        for sube in ham_subeler:
            if isinstance(sube, dict):
                s_id = sube.get("sube_id") or sube.get("id") or sube.get("sube_id AS id")
                formatli_subeler.append({
                    "id": str(s_id) if s_id is not None else "",
                    "sube_adi": str(sube.get("sube_adi", "")),
                    "enlem": float(sube.get("enlem") or 0.0),
                    "boylam": float(sube.get("boylam") or 0.0),
                    "guvenli_yari_cap": int(sube.get("guvenli_yari_cap") or 50)
                })
        return JSONResponse(content={"status": "success", "data": formatli_subeler})
    except Exception as e:
        print(f"Sube listesi cekme hatasi: {e}")
        return JSONResponse(content={"status": "success", "data": []})
        
@app.post("/api/admin/sube-ekle")
async def api_sube_ekle(
    sube_adi: str = Form(...), enlem: str = Form(...),
    boylam: str = Form(...), guvenli_yari_cap: str = Form(...)
):
    basari, mesaj = veritabani.sube_ekle(sube_adi, enlem, boylam, guvenli_yari_cap)
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})

@app.post("/api/admin/sube-sil")
async def api_sube_sil(sube_id: str = Form(...)):
    basari, mesaj = veritabani.sube_sil(sube_id)
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})

@app.post("/api/admin/personel-guncelle")
async def api_personel_guncelle(request: Request):
    form = await request.form()
    try:
        sube_atamalari = sube_atamalarini_oku(form)
    except ValueError as exc:
        return JSONResponse(content={"status": "error", "message": str(exc)})
    tc = str(form.get("tc_kimlik_no", "")).strip()
    if not tc_kimlik_gecerli(tc):
        return JSONResponse(content={"status": "error", "message": "Geçerli 11 haneli TC kimlik numarası zorunludur."})
    ek = form_personel_alanlari(form)
    ek["tc_kimlik_no"] = tc.upper() if tc.upper().startswith("TEST") else tc
    if not all(str(form.get(a, "")).strip() for a in ("isim", "soyisim", "sicil_no", "il", "ilce", "acik_adres")):
        return JSONResponse(content={"status": "error", "message": "Ad, soyad, sicil numarası, il, ilçe ve açık adres zorunludur."})
    if ek.get("cinsiyet") == "Erkek" and not ek.get("askerlik_durumu"):
        return JSONResponse(content={"status": "error", "message": "Erkek personel için askerlik durumu zorunludur."})
    if getattr(form.get("foto"), "filename", ""):
        try:
            ek["foto_mime"], ek["foto_base64"] = await fotograf_hazirla(form.get("foto"))
        except ValueError as exc:
            return JSONResponse(content={"status": "error", "message": str(exc)})
    else:
        dbc = veritabani.baglanti_ac()
        foto_var = dbc.execute("SELECT foto_base64 FROM personeller WHERE id=?", (form.get("p_id", ""),)).fetchone()
        dbc.close()
        if not foto_var or not foto_var[0]:
            return JSONResponse(content={"status": "error", "message": "Personel fotoğrafı zorunludur."})
    basari, mesaj = veritabani.personel_guncelle(
        form.get("p_id", ""), form.get("isim", ""), form.get("soyisim", ""),
        form.get("departman", ""), form.get("maas", "0"),
        form.get("calisma_modeli", "SABİT"), form.get("sicil_no", ""),
        form.get("telefon", ""), form.get("gorev", ""), form.get("sube_id", ""),
        form.get("aktif", "1"), **ek
    )
    if basari:
        basari, sube_mesaji = veritabani.personel_subelerini_ayarla(
            form.get("p_id", ""), form.get("sube_id", ""), sube_atamalari
        )
        if not basari:
            mesaj = "Personel güncellendi ancak şube yetkileri kaydedilemedi: " + sube_mesaji
        else:
            basari, amir_mesaji = veritabani.personel_amir_ata(form.get("p_id", ""), form.get("amir_id", ""))
            if not basari:
                mesaj = "Personel güncellendi ancak ilgili amir kaydedilemedi: " + amir_mesaji
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})

@app.get("/api/admin/firma-ayarlari")
def firma_ayarlari_getir():
    return JSONResponse(content={"status": "success", "data": veritabani.firma_ayarlarini_getir()})

@app.get("/api/admin/ise-gelmeyen-kontrol")
def admin_ise_gelmeyen_kontrol():
    eklenen = veritabani.gelmeyen_personelleri_kontrol_et()
    data = veritabani.duzeltme_talepleri_getir(tumu=False)
    gelmeyen = [x for x in data if str(x.get("talep_turu","")).upper() == "İŞE GELMEDİ"]
    return JSONResponse(content={"status":"success","eklenen":eklenen,"ise_gelmeyen_sayisi":len(gelmeyen),"data":gelmeyen})

@app.get("/api/admin/duzeltme-talepleri")
def admin_duzeltme_talepleri(tumu: int=Query(0)):
    veritabani.gelmeyen_personelleri_kontrol_et()
    return JSONResponse(content={"status":"success","data":veritabani.duzeltme_talepleri_getir(tumu=bool(tumu))})

@app.post("/api/admin/duzeltme-karar")
def admin_duzeltme_karar(talep_id: str=Form(...), karar: str=Form(...), duzeltilmis_zaman: str=Form(""), aciklama: str=Form("")):
    ok,msg=veritabani.duzeltme_talebi_kararla(talep_id,karar,duzeltilmis_zaman,aciklama,"Yönetici")
    return JSONResponse(content={"status":"success" if ok else "error","message":msg})

@app.post("/api/admin/firma-ayarlari")
def firma_ayarlari_guncelle(gec_kalma_kontrolu: str = Form("0"), tolerans_dakika: str = Form("20"), test_modu: str = Form("0")):
    try:
        veritabani.firma_ayarlarini_guncelle(gec_kalma_kontrolu, tolerans_dakika, test_modu)
        return JSONResponse(content={"status": "success", "message": "Firma ayarları kaydedildi."})
    except ValueError:
        return JSONResponse(content={"status": "error", "message": "Tolerans süresi geçersiz."})

@app.post("/api/admin/tum-personel-verilerini-temizle")
def tum_personel_verilerini_temizle(onay_metni: str = Form(...)):
    if onay_metni.strip().upper() != "TÜM PERSONELLERİ SİL":
        return JSONResponse(content={"status": "error", "message": "Güvenlik onay metni yanlış."})
    basari, mesaj = veritabani.tum_personel_verilerini_temizle()
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})

@app.post("/api/admin/sube-guncelle")
async def api_sube_guncelle(
    s_id: str = Form(...), sube_adi: str = Form(...),
    enlem: str = Form(...), boylam: str = Form(...), 
    guvenli_yari_cap: str = Form(...)
):
    basari, mesaj = veritabani.sube_guncelle(s_id, sube_adi, enlem, boylam, guvenli_yari_cap)
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})
@app.post("/api/personel/cihaz-bagla")
async def personel_cihaz_bagla(
    personel_id: str = Form(...), gelen_cihaz_id: str = Form(...)
):
    return JSONResponse(
        status_code=410,
        content={"status": "error", "message": "Eski cihaz bağlama yöntemi kapatıldı."}
    )

@app.post("/api/personel/kurulum")
@limiter.limit("5/minute")
async def personel_kurulum_api(
    request: Request,
    sicil_no: str = Form(...),
    cihaz_id: str = Form(...),
    pin: str = Form(...),
    pin_tekrar: str = Form(""),
    vpn_proxy_aktif: str = Form("0"),
    android_guvenlik_surumu: str = Form("0"),
    kvkk_metin_surumu: str = Form(""),
    kvkk_bilgi_zamani: str = Form("")
):
    personel = veritabani.personel_sicil_ile_getir(sicil_no)
    if not personel:
        return JSONResponse(content={"status": "error", "message": "Sicil numarası bulunamadı."})
    if not personel.get("aktif"):
        return JSONResponse(content={"status": "error", "message": "Personel hesabı pasif."})
    if not personel.get("sube_id"):
        return JSONResponse(content={"status": "error", "message": "Önce personele bir şube atanmalıdır."})

    if kvkk_metin_surumu != KVKK_METIN_SURUMU or not kvkk_bilgi_zamani:
        return JSONResponse(content={"status": "error", "message": "KVKK aydınlatma adımı tamamlanmadan cihaz kurulumu yapılamaz. Uygulamayı kapatıp yeniden açın."})

    if str(vpn_proxy_aktif).lower() in ("1", "true", "evet"):
        veritabani.hata_logu_yaz(personel["id"], "AĞ", "VPN_PROXY", "Kurulum sırasında VPN veya proxy tespit edildi.")
        return JSONResponse(content={"status":"error", "message":"VPN/Proxy tespit edildi. Fake IP/VPN kapatıp tekrar deneyin."})

    mevcut_cihaz = personel.get("cihaz_id")
    if mevcut_cihaz not in (None, "", "EŞLEŞMEDİ", cihaz_id):
        return JSONResponse(content={
            "status": "error",
            "message": "Bu personel başka bir telefona bağlı. Yönetici cihaz kaydını sıfırlamalıdır."
        })

    if not pin.isdigit() or len(pin) != 6:
        return JSONResponse(content={"status": "error", "message": "Şifre 6 rakam olmalıdır."})
    if personel.get("personel_pin_hash"):
        kilitli, kalan = veritabani.personel_pin_kilitli_mi(personel)
        if kilitli:
            return JSONResponse(content={"status":"error", "message":f"Çok sayıda hatalı deneme nedeniyle giriş {kalan} dakika kilitli."})
        if not veritabani.personel_pin_dogrula(personel, pin):
            veritabani.personel_pin_deneme_kaydet(personel["id"], False)
            veritabani.hata_logu_yaz(personel["id"], "GİRİŞ", "PIN_HATALI", "Hatalı personel şifresi girildi.")
            return JSONResponse(content={"status": "error", "message": "Sicil numarası veya şifre hatalı."})
        veritabani.personel_pin_deneme_kaydet(personel["id"], True)
    else:
        if pin != pin_tekrar:
            return JSONResponse(content={"status": "error", "message": "Şifreler eşleşmiyor."})
        if not veritabani.personel_pin_kaydet(personel["id"], pin):
            return JSONResponse(content={"status": "error", "message": "Şifre oluşturulamadı."})

    cihaz_token = secrets.token_urlsafe(32)
    token_hash = hashlib.sha256(cihaz_token.encode()).hexdigest()
    if not veritabani.cihaz_kurulumunu_tamamla(personel["id"], cihaz_id, token_hash):
        return JSONResponse(content={"status": "error", "message": "Cihaz kurulumu tamamlanamadı."})
    veritabani.kvkk_bilgilendirme_kaydet(personel["id"], kvkk_metin_surumu, kvkk_bilgi_zamani)

    return JSONResponse(content={
        "status": "success",
        "message": f"{personel['isim']} {personel['soyisim']} için cihaz kurulumu tamamlandı.",
        "personel": f"{personel['isim']} {personel['soyisim']}",
        "sube": personel["sube_adi"],
        "cihaz_token": cihaz_token
    })

@app.post("/api/personel/kurulum-kontrol")
@limiter.limit("10/minute")
async def personel_kurulum_kontrol(request: Request, sicil_no: str = Form(...), cihaz_id: str = Form(...), vpn_proxy_aktif: str = Form("0")):
    personel = veritabani.personel_sicil_ile_getir(sicil_no)
    if not personel or not personel.get("aktif"):
        return JSONResponse(content={"status": "error", "message": "Aktif personel kaydı bulunamadı."})
    if not personel.get("sube_id"):
        return JSONResponse(content={"status": "error", "message": "Önce personele bir şube atanmalıdır."})
    if str(vpn_proxy_aktif).lower() in ("1", "true", "evet"):
        veritabani.hata_logu_yaz(personel["id"], "AĞ", "VPN_PROXY", "Sicil kontrolünde VPN veya proxy tespit edildi.")
        return JSONResponse(content={"status":"error", "message":"VPN/Proxy tespit edildi. Fake IP/VPN kapatıp tekrar deneyin."})
    if personel.get("cihaz_id") not in (None, "", "EŞLEŞMEDİ", cihaz_id):
        return JSONResponse(content={"status": "error", "message": "Bu personel başka bir telefona bağlı. Yönetici cihaz kaydını sıfırlamalıdır."})
    return JSONResponse(content={
        "status": "success",
        "pin_var": bool(personel.get("personel_pin_hash")),
        "personel": f"{personel['isim']} {personel['soyisim']}"
    })

@app.post("/api/admin/personel-cihaz-sifirla")
async def personel_cihaz_sifirla(personel_id: str = Form(...)):
    basarili = veritabani.personel_cihaz_sadece_sifirla(personel_id)
    return JSONResponse(content={
        "status": "success" if basarili else "error",
        "message": "Yalnızca cihaz eşleştirmesi sıfırlandı. Personel şifresi değiştirilmedi." if basarili else "Personel bulunamadı."
    })

@app.post("/api/admin/personel-sifre-sifirla")
async def personel_sifre_sifirla(personel_id: str = Form(...)):
    basarili = veritabani.personel_sifre_sifirla(personel_id)
    return JSONResponse(content={
        "status": "success" if basarili else "error",
        "message": "Personel şifresi sıfırlandı. Cihaz eşleştirmesi değiştirilmedi." if basarili else "Personel bulunamadı."
    })

@app.post("/api/personel/erisim-talebi")
@limiter.limit("5/minute")
async def personel_erisim_talebi(request: Request, sicil_no: str=Form(...), talep_turu: str=Form(...), aciklama: str=Form("")):
    ok,msg = veritabani.erisim_talebi_olustur_sicil(sicil_no,talep_turu,aciklama)
    return JSONResponse(content={"status":"success" if ok else "error","message":msg})

@app.get("/api/admin/erisim-talepleri")
def admin_erisim_talepleri():
    return JSONResponse(content={"status":"success","data":veritabani.erisim_taleplerini_getir(tumu=False)})

@app.post("/api/admin/erisim-talebi-karar")
async def admin_erisim_talebi_karar(talep_id: str=Form(...), karar: str=Form(...)):
    ok,msg=veritabani.erisim_talebi_kararla(talep_id,karar,"Yönetici")
    return JSONResponse(content={"status":"success" if ok else "error","message":msg})

@app.get("/api/admin/personel-kartlari/{personel_id}")
def personel_kartlari(personel_id: int):
    return {"status": "success", "data": veritabani.personel_kartlarini_getir(personel_id)}

@app.post("/api/admin/personel-kart-ata")
def personel_kart_ata(
    personel_id: str = Form(...), kart_no: str = Form(...),
    kart_turu: str = Form("RFID"), gecerlilik_tarihi: str = Form("")
):
    if kart_turu not in ("RFID", "NFC", "QR"):
        return JSONResponse(content={"status": "error", "message": "Kart türü geçersiz."})
    basari, mesaj = veritabani.kart_ata(personel_id, kart_no, kart_turu, gecerlilik_tarihi)
    return JSONResponse(content={"status": "success" if basari else "error", "message": mesaj})

@app.get("/yonetici-giris", response_class=HTMLResponse)
def yonetici_giris_ekrani():
    dosya_yolu = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), 
        "yonetici_paneli_gelismis.html"
    )
    try:
        with open(dosya_yolu, "r", encoding="utf-8") as f:
            return HTMLResponse(content=f.read())
    except FileNotFoundError:
        raise HTTPException(
            status_code=404, 
            detail="yonetici_paneli_gelismis.html bulunamadı!"
        )

# =========================================================================
# HARİTA SERVİS MOTORU (YATAY TAŞMALAR TAMAMEN ENGELLENDİ)
# =========================================================================
@app.get('/harita-stil.css')
def harita_css_servis():
    css_kodlari = (
        ".leaflet-container{overflow:hidden}.leaflet-control,"
        ".leaflet-pane{z-index:400}.leaflet-top,.leaflet-bottom"
        "{position:absolute;z-index:1000;pointer-events:none}"
        ".leaflet-top{top:0}.leaflet-right{right:0}.leaflet-bottom"
        "{bottom:0}.leaflet-left{left:0}.leaflet-control"
        "{pointer-events:auto;pointer-events:initial}"
        ".leaflet-pane{position:absolute;left:0;top:0}"
        ".leaflet-tile-pane{z-index:200}.leaflet-overlay-pane"
        "{z-index:400}.leaflet-shadow-pane{z-index:500}"
        ".leaflet-marker-pane{z-index:600}.leaflet-tooltip-pane"
        "{z-index:650}.leaflet-popup-pane{z-index:700}"
        ".leaflet-map-pane canvas{position:absolute;left:0;top:0}"
        ".leaflet-map-pane svg{position:absolute;left:0;top:0}"
        ".leaflet-tile{position:absolute;left:0;top:0;"
        "-webkit-user-select:none;-moz-user-select:none;"
        "user-select:none;pointer-events:none}"
        ".leaflet-tile-container{position:absolute;left:0;top:0}"
        ".leaflet-pixelflipped{display:none}.leaflet-marker-icon,"
        ".leaflet-marker-shadow{position:absolute;left:0;top:0;"
        "display:block}.leaflet-container{background:#ddd;"
        "outline-width:0}.leaflet-container a{color:#0078A8}"
        ".leaflet-zoom-animated{-webkit-transform-origin:0 0;"
        "-ms-transform-origin:0 0;transform-origin:0 0}"
        ".leaflet-zoom-anim .leaflet-zoom-animated"
        "{-webkit-transition:-webkit-transform .25s "
        "cubic-bezier(0,0,.25,1);transition:transform .25s "
        "cubic-bezier(0,0,.25,1)}.leaflet-zoom-anim .leaflet-tile,"
        ".leaflet-pan-anim .leaflet-tile{-webkit-transition:none;"
        "transition:none}.leaflet-interactive{cursor:pointer}"
        ".leaflet-grab{cursor:grab}.leaflet-control-zoom"
        "{border-radius:4px;background:#fff;border:2px solid "
        "rgba(0,0,0,0.2)}.leaflet-control-zoom a{width:26px;"
        "height:26px;line-height:26px;display:block;"
        "text-align:center;text-decoration:none;color:black;"
        "font-weight:bold}.leaflet-bar a:hover"
        "{background-color:#f4f4f4}.leaflet-control-attribution"
        "{background:#fff;background:rgba(255,255,255,0.8);"
        "margin:0;padding:0 5px;font-size:11px}"
    )
    return HTMLResponse(
        content=css_kodlari, 
        status_code=200, 
        headers={'Content-Type': 'text/css'}
    )

@app.get('/harita-motoru.js')
def harita_js_servis():
    js_kodlari = (
        "window.L = window.L || {};"
        "L.Map = function(t, e) {"
        "  return {"
        "    setView: function(coords, zoom) {"
        "      var n = document.getElementById('harita');"
        "      var latBox = document.getElementById('s_enlem');"
        "      var lngBox = document.getElementById('s_boylam');"
        "      if (n && latBox && lngBox) {"
        "        var lat = latBox.value;"
        "        var lng = lngBox.value;"
        "        if (lat && lng) {"
        "          n.innerHTML = '<iframe width=\"100%\" "
        "height=\"100%\" style=\"border:0;border-radius:8px;\" "
        "src=\"https://google.com' + lat + ',' + lng + "
        "'&z=' + zoom + '&output=embed\"></iframe>';"
        "        }"
        "      }"
        "      return this;"
        "    }"
        "  };"
        "};"
        "L.map = function(t, e) { return new L.Map(t, e); };"
        "L.tileLayer = function(t, e) { return { addTo: function(m) {} }; };"
        "L.marker = function(t, e) { return { addTo: function(m) { "
        "return { bindPopup: function(x) { return { openPopup: "
        "function() {} }; } }; } }; };"
        "L.circle = function(t, e) { return { addTo: function(m) {} }; };"
    )
    return HTMLResponse(
        content=js_kodlari, 
        status_code=200, 
        headers={'Content-Type': 'text/javascript'}
    )




if __name__ == "__main__":
    veritabani.veritabani_hazirla()
    uvicorn.run(app, host="0.0.0.0", port=int(os.getenv("PORT", "8000")))
